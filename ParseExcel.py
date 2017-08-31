#encoding=utf-8
import re
from copy import deepcopy
from openpyxl import *
from openpyxl.styles import Border, Side, Font
import time
import os
#from Config.GloableData import *
#from Util.FormatTime import date_time_chinese


class ParseExcel(object):
    u"""Excel操作类"""
    def __init__(self, excel_file_path, sheet_name=None, new_file_path=None):
        self.excel_file_path = excel_file_path
        self.new_file_path = new_file_path
        self.workbook = load_workbook(self.excel_file_path)
        if sheet_name:
            self.worksheet = self.workbook.get_sheet_by_name(sheet_name)
        else:
            self.worksheet = self.workbook.active
        self.font = Font(color=None)
        self.colorDict = {"red":'FFFF3030', "green":'FF008B00'}

    def get_sheet_by_name(self, sheet_name):
        # return worksheet
        self.worksheet = self.workbook.get_sheet_by_name(sheet_name)
        return self.worksheet

    def get_all_sheet_names(self):
        return self.workbook.get_sheet_names()

    def get_sheet_by_index(self, sheet_index):
        # return worksheet
        self.worksheet = self.workbook.worksheets[sheet_index]
        return self.worksheet

    def get_sheet_name(self):
        # return sheet_name
        return self.worksheet.title

    def get_max_row_number(self):
        # return max_row_number
        return self.worksheet.max_row

    def get_min_row_number(self):
        # return min_row_number
        return self.worksheet.min_row

    def get_max_col_number(self):
        # return max_col_number
        return self.worksheet.max_column

    def get_min_col_number(self):
        # return min_col_number
        return self.worksheet.min_column

    def get_all_rows(self):
        #return rows
        return self.worksheet.rows

    def get_all_cols(self):
        #return cols
        return self.worksheet.columns

    def get_single_row(self, row_no):
        #return row
        return self.get_all_rows()[row_no - 1]

    def get_single_col(self, col_no):
        #return col
        return self.get_all_cols()[col_no - 1]

    def get_cell(self,row_no,col_no):
        #return cell
        return self.worksheet.cell(row=row_no, column=col_no)

    def get_cell_content(self,row_no,col_no):
        #return cell.value
        return self.get_cell(row_no, col_no).value

    def get_all_rows_content(self):
        # return datas
        datas = []
        for row in self.get_all_rows():
            data = []
            for cell in row:
                value = cell.value
                if value == None:
                    data.append(u'')
                else:
                    data.append(value)
            datas.append(data)
        return datas

    def write_cell_content(self,row_no,col_no,content,font=None):
        #return none
        self.get_cell(row_no, col_no).value = content
        # self.save_file(self.new_file_path)

    def write_cell_current_time(self,row_no,col_no):
        #return none
        self.get_cell(row_no, col_no).value = date_time_chinese()
        self.save_file(self.new_file_path)

    def save_file(self, excel_file_path=None):
        if excel_file_path:
            self.workbook.save(excel_file_path)
        else:
            self.workbook.save(self.excel_file_path)

    def get_data_dict(self):
        u"""文件第一行为key，从第二行往后为value"""
        datas = []
        for row in xrange(2, self.get_max_row_number()+1):
            data = {}
            for col in xrange(1, self.get_max_col_number()+1):
                value = self.get_cell_content(row,col)
                data[self.get_cell_content(1, col)] = value
            datas.append(data)
        return datas

    def get_title(self):
        u"""获取第一行标题"""
        data = []
        for col in xrange(1, self.get_max_col_number()+1):
            data.append(self.get_cell_content(1, col))
        return data

    def get_values_and_actions_dict(self):
        u"""获取数据和操作以字典返回"""
        result = {}
        values, actions = self.get_values_and_actions_arr()
        result['values_title'] = values[0]
        result['actions_title'] = actions[0]
        result['values'] = self.exchange_arrs_to_dicts(values)
        result['actions'] = self.exchange_arrs_to_dicts(actions)
        return result

    def get_values_and_actions_arr(self):
        u"""获取数据和操作以数组返回"""
        values, actions = [], []
        datas = self.get_all_rows_content()
        for id, data in enumerate(datas):
            if list(set(data)) == ['']:
                actions, values = datas[:id], datas[id + 1:]
                break
        return values, actions

    def exchange_arrs_to_dicts(self, arrs):
        u"""将一个二维数组转换成字典数组"""
        results = []
        keys = arrs[0]
        for arr in arrs[1:]:
            result = {}
            for id, key in enumerate(keys):
                if key:
                    result[key] = arr[id]
            results.append(result)
        return results

if __name__ == '__main__':
    # pe = ParseExcel("e:\\test.xlsx", new_file_path="e:\\tset3.xlsx")
    # print pe.get_max_row_number()
    # print pe.get_min_row_number()
    # print pe.get_max_col_number()
    # print pe.get_min_col_number()
    # print pe.get_all_rows()
    # print pe.get_all_cols()
    # print pe.get_cell(1, 1)
    # print pe.get_cell_content(1, 1)
    # for cell in pe.get_single_row(1):
    #     print cell.value,
    # print
    # print pe.get_single_col(1)
    # pe.write_cell_content(3, 6, "哈哈")
    # pe.write_cell_current_time(3, 5)
    pe = ParseExcel(test_data_excel_path, u"联系人")